import openpyxl
wb=openpyxl.load_workbook("..//PANDASPROJECT/Book1.xlsx")
sheet=wb.active
data=sheet['A3'].value
data3=sheet.cell(row=2,column=3)
data1=sheet['A1:A10']
#print(data)
#print(data1)
#print(data3)
#Access all cell in row1
print(sheet.max_row)
print(sheet.max_column)
#for row in sheet.rows:
    #print([data.value for data in row])
for i in range(1,12):
    print(sheet.cell(row=i,column=1).value)
sheet["H2"].value='=SUM(C2:G2)'